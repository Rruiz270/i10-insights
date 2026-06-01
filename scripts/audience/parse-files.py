#!/usr/bin/env python3
"""
Parse the external contact bases into a normalized JSON array on stdout.

Each output record: {email, phone, name, role, municipio, uf, source, segment,
consent, attributes}. A municipality row explodes into several records (mayor /
city-hall / education secretary), each with its own role.

Usage:
  python3 parse-files.py <geral_br.xlsx> <paraiba.html> <brasil_edu.html> > out.json
Any path may be omitted with "-" to skip that source.
"""
import sys, re, json, html as htmllib

UF_FROM_DOMAIN = re.compile(r"\.([a-z]{2})\.gov\.br", re.I)


def clean_email(v):
    if not v:
        return None
    v = str(v).strip().lower()
    return v if "@" in v and "." in v.split("@")[-1] else None


def clean_phone(v):
    if not v:
        return None
    digits = re.sub(r"\D", "", str(v))
    return digits if len(digits) >= 10 else None


def rec(email, phone, name, role, municipio, uf, source, segment, consent, **attrs):
    email, phone = clean_email(email), clean_phone(phone)
    if not email and not phone:
        return None
    return {
        "email": email, "phone": phone,
        "name": (name or "").strip() or None,
        "role": role, "municipio": (municipio or "").strip() or None,
        "uf": (uf or "").strip().upper() or None,
        "source": source, "segment": segment, "consent": consent,
        "attributes": {k: v for k, v in attrs.items() if v not in (None, "", ".")},
    }


# ── 1. GERAL BR (mayors xlsx) ────────────────────────────────────────
def parse_mayors(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [str(c).strip().lower() if c is not None else "" for c in next(rows)]
    idx = {h: i for i, h in enumerate(header)}

    def col(r, *names):
        for n in names:
            if n in idx and idx[n] < len(r):
                return r[idx[n]]
        return None

    out = []
    for r in rows:
        mun = str(col(r, "nome_municipio") or "").strip()
        uf = col(r, "uf") or (mun.split("/")[-1] if "/" in mun else None)
        mun = mun.split("/")[0].strip()
        out.append(rec(
            col(r, "e mail", "email"), col(r, "celular1"), col(r, "nome"),
            "prefeito", mun, uf, "file-prefeitos-br", "prefeito-base", None,
            partido=col(r, "partido"), populacao=col(r, "população", "populacao"),
            reeleito=col(r, "reeleito"), apelido=col(r, "apelido"),
            genero=col(r, "gênero", "genero"), celular2=clean_phone(col(r, "celular2")),
        ))
    return [o for o in out if o]


# ── 2. Paraíba (clean HTML table) ────────────────────────────────────
def strip_tags(s):
    return htmllib.unescape(re.sub(r"<[^>]+>", "", s or "")).strip()


def cell_email(td):
    m = re.search(r"mailto:([^\"'>]+)", td)
    return m.group(1) if m else strip_tags(td)


def cell_phone(td):
    m = re.search(r"tel:([+\d]+)", td)
    return m.group(1) if m else strip_tags(td)


def parse_paraiba(path):
    doc = open(path, encoding="utf-8", errors="ignore").read()
    body = doc[doc.find("<tbody"):] if "<tbody" in doc else doc
    out = []
    for tr in re.findall(r"<tr[^>]*>(.*?)</tr>", body, re.S | re.I):
        tds = re.findall(r"<td[^>]*>(.*?)</td>", tr, re.S | re.I)
        if len(tds) < 9:
            continue
        mun, pref, part = strip_tags(tds[0]), strip_tags(tds[1]), strip_tags(tds[2])
        site = re.search(r"href=\"(https?://[^\"]+)", tds[9]) if len(tds) > 9 else None
        site = site.group(1) if site else None
        triples = [
            ("prefeito", "prefeito-base", cell_email(tds[3]), cell_phone(tds[6]), pref),
            ("prefeitura", "prefeitura", cell_email(tds[4]), cell_phone(tds[7]), f"{mun} (prefeitura)"),
            ("educacao", "secretaria-educacao", cell_email(tds[5]), cell_phone(tds[8]), f"{mun} (educação)"),
        ]
        for role, seg, em, ph, nm in triples:
            out.append(rec(em, ph, nm, role, mun, "PB", "file-paraiba", seg, None,
                           partido=part, site=site))
    return [o for o in out if o]


# ── 3. Brasil Prefeitos+Educação (embedded JSON array) ───────────────
def extract_json_arrays(doc):
    """The page paginates data into many [{...}] arrays — return them all."""
    arrays, i = [], 0
    while True:
        start = doc.find("[{", i)
        if start < 0:
            break
        depth, in_str, esc, end = 0, False, False, -1
        for j in range(start, len(doc)):
            c = doc[j]
            if in_str:
                if esc:
                    esc = False
                elif c == "\\":
                    esc = True
                elif c == '"':
                    in_str = False
            else:
                if c == '"':
                    in_str = True
                elif c == "[":
                    depth += 1
                elif c == "]":
                    depth -= 1
                    if depth == 0:
                        end = j + 1
                        break
        if end < 0:
            break
        arrays.append(doc[start:end])
        i = end
    return arrays


def parse_brasil(path):
    doc = open(path, encoding="utf-8", errors="ignore").read()
    data = []
    for arr in extract_json_arrays(doc):
        try:
            data += json.loads(arr)
        except Exception:
            pass
    out = []
    for d in data:
        mun = (d.get("m") or "").strip()
        dom = d.get("d") or ""
        uf = (UF_FROM_DOMAIN.search(dom).group(1) if UF_FROM_DOMAIN.search(dom) else None)
        pref, part, site = d.get("p"), d.get("pt"), d.get("s")
        pairs = [
            ("prefeitura", "gabinete", d.get("ge"), d.get("gp"), pref),
            ("educacao", "secretaria-educacao", d.get("ee"), d.get("ep"), f"{mun} (educação)"),
        ]
        for role, seg, em, ph, nm in pairs:
            out.append(rec(em, ph, nm, role, mun, uf, "file-brasil-edu", seg, None,
                           partido=part, site=site, dominio=dom))
    return [o for o in out if o]


def main():
    args = sys.argv[1:]
    parsers = [parse_mayors, parse_paraiba, parse_brasil]
    records = []
    for path, fn in zip(args, parsers):
        if path and path != "-":
            try:
                got = fn(path)
                records += got
                print(f"  {fn.__name__}: {len(got)} contatos", file=sys.stderr)
            except Exception as e:
                print(f"  {fn.__name__} FALHOU: {e}", file=sys.stderr)
    json.dump(records, sys.stdout, ensure_ascii=False)


if __name__ == "__main__":
    main()
